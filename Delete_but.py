from PyQt5.Qt import *
import openpyxl
import Warning_window
class Delete_button(QPushButton):
    def __init__(self,parent):
        super().__init__(parent)
        self.Gun_name = ''
        self.Account = ''
        self.clicked.connect(self.Delete_parent)
    def Delete_parent(self):
        try:
            wb = openpyxl.load_workbook('./UserData/' + self.Account + '/Weapon.xlsx')
            ws = wb["Weapon"]
            row = 1
            while True:
                if str(ws.cell(row=row, column=1).value) != self.Gun_name:
                    row += 1
                    if row > 6:
                        break
                else:
                    if row == 1:
                        ws['G2'] = int(int(ws.cell(row=1,column=7).value) - 1)
                        ws.delete_rows(row)
                    else:
                        ws.cell(row=1, column=7).value = int(int(ws.cell(row=1,column=7).value) - 1)
                        ws.delete_rows(row)
                    wb.save("./UserData/" + self.Account + "/Weapon.xlsx")
                    self.parent().deleteLater()
                    self.parent().parent().parent().parent().parent().Add_Weapon.AddWeapon.check_list.remove(self.Gun_name)
                    break
        except:
            self.er = Warning_window.Error()
            self.er.show()